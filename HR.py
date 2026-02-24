# pyinstaller --onefile --windowed --icon=HR_icon.ico --add-data "HR_icon.ico:." --add-data "fb_include/fbclient.dll:." --add-data "fb_include/firebird.msg:." --add-data "fb_include/icudt30.dll:." HR.py
import sys
import os
import datetime
import fdb
import re
from PyQt5 import QtWidgets, QtCore, QtGui
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import base64

SECRET_KEY = "HR_secret_key_2024"


def encrypt_password(password: str) -> str:
    data = password.encode("utf-8")
    key = SECRET_KEY.encode("utf-8")

    encrypted = bytes([b ^ key[i % len(key)] for i, b in enumerate(data)])
    return base64.b64encode(encrypted).decode("utf-8")


def decrypt_password(encoded: str) -> str:
    try:
        data = base64.b64decode(encoded.encode("utf-8"))
        key = SECRET_KEY.encode("utf-8")

        decrypted = bytes([b ^ key[i % len(key)] for i, b in enumerate(data)])
        return decrypted.decode("utf-8")
    except Exception:
        return ""


# ================= MODEL =================

class EmployeesModel(QtCore.QAbstractTableModel):

    headers = [
        "№",
        "ФИО",
        "Дата приема",
        "Стаж (лет)",
        "До очередной выслуги",
        "Примечание"
    ]

    def __init__(self, connection):
        super().__init__()
        self.conn = connection
        self.cur = connection.cursor()
        self.load()

    # ---------- Загрузка данных ----------
    def load(self):
        self.cur.execute("SELECT id, fio, hire_date, note FROM employees ORDER BY id")
        self.rows = self.cur.fetchall()

    # ---------- Размеры ----------
    def rowCount(self, parent=None):
        return len(self.rows)

    def columnCount(self, parent=None):
        return 6

    # ---------- Отображение ----------
    def data(self, index, role):

        if not index.isValid():
            return None

        emp_id, fio, hire_date, note = self.rows[index.row()]
        col = index.column()

        # ===== Отображение текста =====
        if role == QtCore.Qt.DisplayRole or role == QtCore.Qt.EditRole:

            if col == 0:
                return index.row() + 1

            if col == 1:
                return fio

            if col == 2 and hire_date:
                return hire_date.strftime("%d.%m.%Y")

            if col == 3 and hire_date:
                return self.calculate_experience(hire_date)

            if col == 4 and hire_date:
                return self.next_milestone(hire_date)

            if col == 5:
                return note

        # ===== Цвет юбилея =====
        if role == QtCore.Qt.BackgroundRole and hire_date:

            settings = QtCore.QSettings("MyCompany", "HRApp")
            years = self.calculate_experience(hire_date)

            # 1️⃣ Проверяем ближайший юбилей (приоритет)
            if settings.value("highlight/month_enabled", False, type=bool):

                months_limit = settings.value("highlight/month_value", 6, type=int)

                next_date = self.get_next_milestone_date(hire_date)
                if next_date:
                    today = datetime.date.today()
                    delta_months = (next_date.year - today.year) * 12 + \
                                   (next_date.month - today.month)

                    if 0 <= delta_months <= months_limit:
                        color = settings.value(
                            "highlight/upcoming_color",
                            "#ff8a80"
                        )
                        return QtGui.QColor(color)

            # 2️⃣ Обычные юбилеи
            milestones = [15, 20, 25, 30]

            reached = None
            for m in milestones:
                if years >= m:
                    reached = m

            if reached:
                color = settings.value(f"highlight/{reached}", None)
                if color:
                    return QtGui.QColor(color)

        return None

    def get_next_milestone_date(self, hire_date):

        milestones = [15, 20, 25, 30]
        years = self.calculate_experience(hire_date)

        for m in milestones:
            if years < m:
                return hire_date.replace(year=hire_date.year + m)

        return None

    # ---------- Заголовки ----------
    def headerData(self, section, orientation, role):
        if role == QtCore.Qt.DisplayRole and orientation == QtCore.Qt.Horizontal:
            return self.headers[section]
        return None

    # ---------- Редактирование ----------
    def flags(self, index):
        if index.column() in (1, 2, 5):
            return QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsEditable
        return QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled

    def setData(self, index, value, role):

        if role != QtCore.Qt.EditRole:
            return False

        row = index.row()
        emp_id, fio, hire_date, note = self.rows[row]

        if index.column() == 1:
            fio = value


        elif index.column() == 2:
            if not value:
                hire_date = None
            else:
                try:
                    hire_date = datetime.datetime.strptime(
                        value, "%d.%m.%Y"
                    ).date()
                except:
                    return False

        elif index.column() == 5:
            note = value

        self.cur.execute("""
            UPDATE employees
            SET fio=?, hire_date=?, note=?
            WHERE id=?
        """, (fio, hire_date, note, emp_id))

        self.conn.commit()

        self.rows[row] = (emp_id, fio, hire_date, note)

        self.dataChanged.emit(
            self.index(row, 0),
            self.index(row, 5)
        )

        return True

    # ---------- Добавление ----------
    def add_employee(self):
        self.cur.execute(
            "INSERT INTO employees (fio, hire_date, note) VALUES (?, ?, ?)",
            ("Новый сотрудник", None, "")
        )
        self.conn.commit()
        self.layoutAboutToBeChanged.emit()
        self.load()
        self.layoutChanged.emit()

    # ---------- Расчет ----------
    def calculate_experience(self, hire_date):
        today = datetime.date.today()
        return today.year - hire_date.year - (
            (today.month, today.day) < (hire_date.month, hire_date.day)
        )

    def next_milestone(self, hire_date):
        milestones = [15, 20, 25, 30, 35, 40, 50, 55, 60]
        years = self.calculate_experience(hire_date)

        for m in milestones:
            if years < m:
                d = hire_date.replace(year=hire_date.year + m)
                return f"{d.strftime('%d.%m.%Y')} — будет {m} лет"

        return "Более 30 лет"

    # ---------- автообновление данных ----------
    def refresh_experience(self):
        if not self.rows:
            return

        top_left = self.index(0, 0)
        bottom_right = self.index(len(self.rows) - 1, 5)

        self.dataChanged.emit(top_left, bottom_right)

    def delete_employee(self, source_row):

        emp_id, fio, hire_date, note = self.rows[source_row]

        self.cur.execute(
            "DELETE FROM employees WHERE id = ?",
            (emp_id,)
        )
        self.conn.commit()

        self.layoutAboutToBeChanged.emit()
        self.load()
        self.layoutChanged.emit()

# ================= DATE DELEGATE =================

class DateDelegate(QtWidgets.QStyledItemDelegate):

    def createEditor(self, parent, option, index):
        editor = QtWidgets.QLineEdit(parent)
        editor.setPlaceholderText("дд.мм.гггг")
        return editor

    def setEditorData(self, editor, index):
        text = index.data(QtCore.Qt.DisplayRole)
        editor.setText(text if text else "")

    def setModelData(self, editor, model, index):
        text = editor.text().strip()

        # Пустая дата
        if not text:
            model.setData(index, "", QtCore.Qt.EditRole)
            return

        parsed = self.parse_date(text)

        if parsed:
            editor.setStyleSheet("")
            model.setData(
                index,
                parsed.strftime("%d.%m.%Y"),
                QtCore.Qt.EditRole
            )
        else:
            # Подсветка ошибки (без всплывающего окна)
            editor.setStyleSheet("border: 1px solid red;")

    # -------- Умный парсинг --------

    def parse_date(self, text):

        text = text.lower().strip()

        # Убираем слово "год"
        text = re.sub(r"\bгод\b", "", text)

        # ---------------- 1️⃣ Цифровые форматы ----------------

        # 01022015
        match = re.search(r"\b(\d{2})(\d{2})(\d{4})\b", text)
        if match:
            try:
                return datetime.date(
                    int(match.group(3)),
                    int(match.group(2)),
                    int(match.group(1))
                )
            except:
                pass

        # 01.02.2015 / 01-02-2015 / 01/02/2015
        match = re.search(
            r"\b(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})\b",
            text
        )
        if match:
            day = int(match.group(1))
            month = int(match.group(2))
            year = int(match.group(3))

            if year < 100:
                year += 2000

            try:
                return datetime.date(year, month, day)
            except:
                pass

        # ---------------- 2️⃣ Форматы с названием месяца ----------------

        months = {
            "январ": 1, "феврал": 2, "март": 3, "апрел": 4,
            "мая": 5, "май": 5, "июн": 6, "июл": 7,
            "август": 8, "сентябр": 9, "октябр": 10,
            "ноябр": 11, "декабр": 12,

            # английские
            "january": 1, "february": 2, "march": 3,
            "april": 4, "may": 5, "june": 6,
            "july": 7, "august": 8,
            "september": 9, "october": 10,
            "november": 11, "december": 12
        }

        match = re.search(
            r"\b(\d{1,2})\s+([а-яa-z]+)\s+(\d{2,4})\b",
            text
        )

        if match:
            day = int(match.group(1))
            month_text = match.group(2)
            year = int(match.group(3))

            if year < 100:
                year += 2000

            for key in months:
                if key in month_text:
                    try:
                        return datetime.date(year, months[key], day)
                    except:
                        pass

        return None


# ================= WINDOW =================
class DbSettingsDialog(QtWidgets.QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки подключения к БД")
        self.resize(400, 250)

        layout = QtWidgets.QFormLayout(self)

        self.host_edit = QtWidgets.QLineEdit()
        self.path_edit = QtWidgets.QLineEdit()
        self.user_edit = QtWidgets.QLineEdit()
        self.pass_edit = QtWidgets.QLineEdit()
        self.pass_edit.setEchoMode(QtWidgets.QLineEdit.Password)
        self.charset_edit = QtWidgets.QLineEdit("WIN1251")
        self.port_edit = QtWidgets.QLineEdit("3050")

        layout.addRow("Хост:", self.host_edit)
        layout.addRow("Путь к БД:", self.path_edit)
        layout.addRow("Пользователь:", self.user_edit)
        layout.addRow("Пароль:", self.pass_edit)
        layout.addRow("Кодировка:", self.charset_edit)
        layout.addRow("Порт:", self.port_edit)


        self.test_btn = QtWidgets.QPushButton("Проверить соединение")
        self.save_btn = QtWidgets.QPushButton("Сохранить")

        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.addWidget(self.test_btn)
        btn_layout.addWidget(self.save_btn)
        layout.addRow(btn_layout)

        self.test_btn.clicked.connect(self.test_connection)
        self.save_btn.clicked.connect(self.accept)

        self.load_settings()

    def load_settings(self):
        settings = QtCore.QSettings("MyCompany", "HRApp")
        self.host_edit.setText(settings.value("db/host", ""))
        self.path_edit.setText(settings.value("db/path", ""))
        self.user_edit.setText(settings.value("db/user", ""))

        encrypted = settings.value("db/password", "")
        self.pass_edit.setText(decrypt_password(encrypted))

        self.charset_edit.setText(settings.value("db/charset", "WIN1251"))
        self.port_edit.setText(settings.value("db/port", "3050"))

    def save_settings(self):
        settings = QtCore.QSettings("MyCompany", "HRApp")
        settings.setValue("db/host", self.host_edit.text())
        settings.setValue("db/path", self.path_edit.text())
        settings.setValue("db/user", self.user_edit.text())

        settings.setValue(
            "db/password",
            encrypt_password(self.pass_edit.text())
        )

        settings.setValue("db/charset", self.charset_edit.text())
        settings.setValue("db/port", self.port_edit.text())

    def test_connection(self):
        try:
            fdb.connect(
                dsn=f"{self.host_edit.text()}:{self.path_edit.text()}",
                user=self.user_edit.text(),
                password=self.pass_edit.text(),
                charset=self.charset_edit.text(),
                port=int(self.port_edit.text())
            )
            QtWidgets.QMessageBox.information(self, "Успех", "Подключение успешно!")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", str(e))

    def accept(self):
        self.save_settings()
        super().accept()


class ExperienceApp(QtWidgets.QMainWindow):

    def __init__(self):
        super().__init__()

        icon_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "HR_icon.ico"
        )

        if os.path.exists(icon_path):
            self.setWindowIcon(QtGui.QIcon(icon_path))

        self.conn = None
        self.init_ui()
        self.connect_to_database()

    def init_ui(self):

        self.setWindowTitle("Учёт выслуги судей")
        self.resize(1200, 650)

        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        layout = QtWidgets.QVBoxLayout(central)

        # ================= Поиск =================
        self.search_edit = QtWidgets.QLineEdit()

        self.search_edit.setPlaceholderText("Поиск по всем столбцам...")
        layout.addWidget(self.search_edit)

        # ================= Таблица =================
        # ВАЖНО: таблица создаётся БЕЗ модели
        self.table = QtWidgets.QTableView()
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.Stretch
        )
        layout.addWidget(self.table)

        # ================= Кнопка =================
        self.add_btn = QtWidgets.QPushButton("Добавить сотрудника")
        layout.addWidget(self.add_btn)

        # ================= Меню =================
        menubar = self.menuBar()

        settings_menu = menubar.addMenu("Настройки")

        db_action = QtWidgets.QAction("Подключение к БД", self)
        db_action.triggered.connect(self.open_settings)
        settings_menu.addAction(db_action)

        highlight_action = QtWidgets.QAction("Выделение цветом", self)
        highlight_action.triggered.connect(self.open_highlight_settings)
        settings_menu.addAction(highlight_action)

        # 👇 НОВОЕ МЕНЮ СПРАВА
        export_menu = menubar.addMenu("Экспорт данных")

        export_action = QtWidgets.QAction("В Excel", self)
        export_action.triggered.connect(self.export_filtered_to_excel)
        export_menu.addAction(export_action)

    def export_filtered_to_excel(self):

        if not hasattr(self, "proxy"):
            QtWidgets.QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта.")
            return

        documents_path = QtCore.QStandardPaths.writableLocation(
            QtCore.QStandardPaths.DocumentsLocation
        )

        if not documents_path:
            documents_path = QtCore.QCoreApplication.applicationDirPath()

        today_str = datetime.date.today().strftime("%Y-%m-%d")
        file_path = f"{documents_path}/Список_сотрудников_{today_str}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        headers = self.model.headers
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in range(self.proxy.rowCount()):

            row_data = []
            row_color = None

            for col in range(self.proxy.columnCount()):
                index = self.proxy.index(row, col)
                value = index.data(QtCore.Qt.DisplayRole)
                row_data.append(value)

                # Берём цвет из BackgroundRole
                if col == 0:
                    bg = index.data(QtCore.Qt.BackgroundRole)
                    if isinstance(bg, QtGui.QColor):
                        row_color = bg.name().replace("#", "")

            ws.append(row_data)

            excel_row_number = ws.max_row

            if row_color:
                fill = PatternFill(
                    start_color=row_color,
                    end_color=row_color,
                    fill_type="solid"
                )

                for cell in ws[excel_row_number]:
                    cell.fill = fill

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(file_path)

        QtGui.QDesktopServices.openUrl(
            QtCore.QUrl.fromLocalFile(file_path)
        )

    def open_highlight_settings(self):
        dialog = HighlightSettingsDialog(self)
        if dialog.exec_():
            if hasattr(self, "model"):
                self.model.refresh_experience()

    def open_settings(self):
        dialog = DbSettingsDialog(self)
        if dialog.exec_():
            self.connect_to_database()

    def connect_to_database(self):

        settings = QtCore.QSettings("MyCompany", "HRApp")

        host = settings.value("db/host", "192.168.0.250")
        path = settings.value("db/path", r"c:\invent\HR.FDB")
        user = settings.value("db/user", "sysdba")

        encrypted = settings.value("db/password", encrypt_password("m"))
        password = decrypt_password(encrypted)

        charset = settings.value("db/charset", "WIN1251")
        port = settings.value("db/port", 3050, type=int)

        if not host or not path:
            self.open_settings()
            return

        try:
            self.conn = fdb.connect(
                dsn=f"{host}:{path}",
                user=user,
                password=password,
                charset=charset,
                port=port
            )
            self.init_model()

        except Exception:
            QtWidgets.QMessageBox.warning(
                self,
                "Ошибка подключения",
                "Не удалось подключиться к базе данных."
            )
            self.open_settings()

    def init_model(self):

        self.model = EmployeesModel(self.conn)
        self.model.refresh_experience()

        self.proxy = QtCore.QSortFilterProxyModel()
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.proxy.setFilterKeyColumn(-1)

        self.search_edit.textChanged.connect(self.proxy.setFilterFixedString)

        self.table.setModel(self.proxy)
        self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.open_context_menu)
        self.table.setItemDelegateForColumn(2, DateDelegate())
        self.table.setSortingEnabled(True)

        self.add_btn.clicked.connect(self.model.add_employee)

    def open_context_menu(self, position):

        index = self.table.indexAt(position)

        if not index.isValid():
            return

        menu = QtWidgets.QMenu()

        delete_action = menu.addAction("Удалить запись")

        action = menu.exec_(self.table.viewport().mapToGlobal(position))

        if action == delete_action:

            reply = QtWidgets.QMessageBox.question(
                self,
                "Подтверждение",
                "Вы действительно хотите удалить выбранную запись?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
            )

            if reply == QtWidgets.QMessageBox.Yes:
                # Переводим proxy row в source row
                source_index = self.proxy.mapToSource(index)
                source_row = source_index.row()

                self.model.delete_employee(source_row)


class HighlightSettingsDialog(QtWidgets.QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки выделения цветом")
        self.resize(400, 300)

        layout = QtWidgets.QFormLayout(self)

        self.colors = {}
        self.buttons = {}

        self.milestones = [15, 20, 25, 30]

        for m in self.milestones:
            btn = QtWidgets.QPushButton()
            btn.clicked.connect(lambda _, x=m: self.choose_color(x))
            layout.addRow(f"{m} лет:", btn)
            self.buttons[m] = btn

        # Галка месяцев
        self.month_checkbox = QtWidgets.QCheckBox(
            "Выделять если до юбилея осталось менее"
        )

        # Цвет для ближайшего юбилея
        self.upcoming_color_btn = QtWidgets.QPushButton()
        layout.addRow("Цвет ближайшего юбилея:", self.upcoming_color_btn)
        self.upcoming_color_btn.clicked.connect(self.choose_upcoming_color)

        self.month_spin = QtWidgets.QSpinBox()
        self.month_spin.setRange(1, 24)
        self.month_spin.setValue(6)

        month_layout = QtWidgets.QHBoxLayout()
        month_layout.addWidget(self.month_checkbox)
        month_layout.addWidget(self.month_spin)
        month_layout.addWidget(QtWidgets.QLabel("месяцев"))

        layout.addRow(month_layout)

        self.save_btn = QtWidgets.QPushButton("Сохранить")
        layout.addRow(self.save_btn)

        self.save_btn.clicked.connect(self.accept)

        self.load_settings()

    def choose_upcoming_color(self):
        current_color = QtGui.QColor(self.upcoming_color)
        color = QtWidgets.QColorDialog.getColor(current_color, self)

        if color.isValid():
            self.upcoming_color = color.name()
            self.upcoming_color_btn.setStyleSheet(
                f"background-color:{self.upcoming_color}"
            )
    def load_settings(self):
        settings = QtCore.QSettings("MyCompany", "HRApp")

        default_colors = {
            15: "#e6ffcb",
            20: "#d2feff",
            25: "#ebefff",
            30: "#fff4bc"
        }

        for m in self.milestones:
            color = settings.value(f"highlight/{m}", default_colors[m])
            self.colors[m] = color
            self.buttons[m].setStyleSheet(f"background-color:{color}")

        self.month_checkbox.setChecked(
            settings.value("highlight/month_enabled", True, type=bool)
        )
        self.month_spin.setValue(
            settings.value("highlight/month_value", 6, type=int)
        )
        self.upcoming_color = settings.value(
            "highlight/upcoming_color",
            "#ffb0a6"
        )

        self.upcoming_color_btn.setStyleSheet(
            f"background-color:{self.upcoming_color}"
        )

    def choose_color(self, milestone):
        current_color = QtGui.QColor(self.colors.get(milestone, "#ffffff"))
        color = QtWidgets.QColorDialog.getColor(current_color, self)

        if color.isValid():
            self.colors[milestone] = color.name()
            self.buttons[milestone].setStyleSheet(
                f"background-color:{color.name()}"
            )

    def accept(self):
        settings = QtCore.QSettings("MyCompany", "HRApp")

        for m, color in self.colors.items():
            settings.setValue(f"highlight/{m}", color)

        settings.setValue("highlight/month_enabled",
                          self.month_checkbox.isChecked())
        settings.setValue("highlight/month_value",
                          self.month_spin.value())
        settings.setValue(
            "highlight/upcoming_color",
            self.upcoming_color
        )

        super().accept()


# ================= RUN =================

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)

    icon_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "HR_icon.ico"
    )

    if os.path.exists(icon_path):
        app.setWindowIcon(QtGui.QIcon(icon_path))
    window = ExperienceApp()
    window.show()
    sys.exit(app.exec_())